import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from fpdf import FPDF

class ExcelToPdfConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel to PDF Converter")

        # Frame to hold widgets
        self.frame = tk.Frame(self.root, padx=10, pady=10)
        self.frame.pack()

        # Labels
        self.label_excel = tk.Label(self.frame, text="Excel File:")
        self.label_excel.grid(row=0, column=0, sticky="w")

        self.label_output = tk.Label(self.frame, text="Output Folder:")
        self.label_output.grid(row=1, column=0, sticky="w")

        # Entry widgets
        self.entry_excel = tk.Entry(self.frame, width=50)
        self.entry_excel.grid(row=0, column=1, padx=5, pady=5)

        self.entry_output = tk.Entry(self.frame, width=50)
        self.entry_output.grid(row=1, column=1, padx=5, pady=5)

        # Buttons
        self.button_browse_excel = tk.Button(self.frame, text="Browse...", command=self.browse_excel)
        self.button_browse_excel.grid(row=0, column=2, padx=5, pady=5)

        self.button_browse_output = tk.Button(self.frame, text="Browse...", command=self.browse_output)
        self.button_browse_output.grid(row=1, column=2, padx=5, pady=5)

        self.button_convert = tk.Button(self.frame, text="Convert", command=self.convert_excel_to_pdf)
        self.button_convert.grid(row=2, column=1, pady=10)

    def browse_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if file_path:
            self.entry_excel.delete(0, tk.END)
            self.entry_excel.insert(0, file_path)

    def browse_output(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.entry_output.delete(0, tk.END)
            self.entry_output.insert(0, folder_path)

    def convert_excel_to_pdf(self):
        input_excel = self.entry_excel.get()
        output_folder = self.entry_output.get()

        if not input_excel or not output_folder:
            messagebox.showerror("Error", "Please select both Excel file and Output folder.")
            return

        try:
            # Load the Excel workbook
            wb = load_workbook(filename=input_excel, read_only=True)

            # Iterate through each sheet in the workbook
            for sheet_name in wb.sheetnames:
                pdf = FPDF()
                pdf.set_auto_page_break(auto=True, margin=15)
                pdf.add_page()
                pdf.set_font("Arial", size=12)

                # Get the specific sheet by name
                ws = wb[sheet_name]

                # Extract data from the worksheet and add it to the PDF
                for row in ws.iter_rows(values_only=True):
                    # Join row values into a string separated by tabs
                    text = "\t".join(map(str, row))
                    pdf.cell(200, 10, txt=text, ln=True)

                # Save PDF to file
                output_filename = os.path.join(output_folder, f"{sheet_name}.pdf")
                pdf.output(output_filename)

            messagebox.showinfo("Success", "Conversion completed successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

def main():
    root = tk.Tk()
    app = ExcelToPdfConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()
